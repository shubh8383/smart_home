from os import link
import time
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from selenium.webdriver.common.keys import Keys
import xlrd
import openpyxl
from selenium.webdriver.chrome.options import Options
import pyautogui
import re

options = Options()
options.add_argument("--headless")
driver=webdriver.Chrome(executable_path="/home/zec/Desktop/selenium/chromedriver.exe", options=options)
Link="https://www.google.com/"
driver.get(Link)
driver.maximize_window()



def search_box(city, country, driver):
    time.sleep(2)
    t=driver.find_element(By.CLASS_NAME, "gLFyf").send_keys("smart homes in "+ city+' ' + country)
    time.sleep(2)
    driver.find_element(By.CLASS_NAME, "gLFyf").send_keys(Keys.ENTER)
    time.sleep(2)
    try:
        driver.find_elements(By.CLASS_NAME, "rllt__details")[1].click()
    except: pass
    try:
        driver.find_element(By.CLASS_NAME, "tiS4rf.Q2MMlc").click()
    except:
        pass
    #to scrap all the websites inside the google map
    get_details(city, country, driver)
    #driver.find_element(By.CLASS_NAME, 'axGQJc').click()
    #to scrap all the websites on the home pages
    #main_page()
    time.sleep(2)

def next_page(driver):
    #to get click on next page 
    driver.find_element(By.LINK_TEXT, 'Next').click()
    time.sleep(5)





def get_details(city, country, driver):
    try:
        print("page")
        company_details=[]
        company_name=driver.find_elements(By.CLASS_NAME, 'dbg0pd')
        for i in company_name:
            i.click()
            time.sleep(5)
            name=i.text
            try:
                web=driver.find_element(By.CLASS_NAME, 'dHS6jb')
                webs=web.get_attribute('href')
            except:
                webs=None
            try:
                addres=driver.find_element(By.CLASS_NAME, 'zloOqf.PZPZlf').text
            except:
                addres=None
            if addres is not None:
                Address=re.sub("Address:", "", addres)
            try:
                num=driver.find_element(By.CLASS_NAME, 'LrzXr.zdqRlf.kno-fv').text
            except:
                num=None
            sub_list=[city, country, name, webs, Address, num]
            print(sub_list)
            company_details.append(sub_list)

        df = pd.DataFrame(company_details)
        df.to_csv('USA_company_list.csv', mode='a', index=False, header=False)
        next_page(driver)
        get_details(city, country, driver)

    except: 
        return driver.get(Link)

def enter_proxy_auth(proxy_username, proxy_password):
    time.sleep(2)
    pyautogui.typewrite(proxy_username)
    pyautogui.press('tab')
    pyautogui.typewrite(proxy_password)
    pyautogui.press('enter')

def create_proxy_driver():
    try:
        options = Options()
        options.add_argument("--headless")
        options.add_argument(f'--proxy-server={"45.13.184.21:6582"}')
        driver = webdriver.Chrome('/home/zec/Desktop/selenium/chromedriver.exe' ,options=options)
        driver.get(Link)
        time.sleep(2)
        username, password= "xrefeshw", "ztiuh2rsszox"
        enter_proxy_auth(username, password)
        time.sleep(5)
        sheet_call(driver)
        return driver
    except: pass


wrkbk = openpyxl.load_workbook("/home/zec/Desktop/smart_home/Smart Scrapping.xlsx")
  
sh = wrkbk.active
wb = openpyxl.load_workbook("/home/zec/Desktop/smart_home/USA.xlsx")
sheets = wb.sheetnames
ws = wb[sheets[1]]
def sheet_call(driver):
    # iterate through excel and display data
    for i in range(1665, ws.max_row+1):
        print("\n")
        print("Row ", i, " data :", ws.max_row+1)
        l=[]
        for j in range(1, ws.max_column+1):
            cell_obj = ws.cell(row=i, column=j)
            print(cell_obj.value, end=" ")
            l.append(cell_obj.value)
        search_box(l[0], l[1], driver)

sheet_call(driver)

# creating dataframe
# importing excel file
#df = pd.read_excel('/home/zec/Desktop/smart_home/Smart Scrapping.xlsx')  
  
# load excel with its path

#search_box(CITY='san francisco')